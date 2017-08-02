#!/usr/bin/python
# -*- coding:utf-8 -*-

from PIL import Image
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Style, PatternFill, Color

wb = Workbook()
ws = wb.get_active_sheet()

im = Image.open('./images/myself_sm.jpg')
w, h = im.size
pix = im.load()
for row in range(1, h):
    for col in range(1, w):
        cell = ws.cell(column=col, row=row, value='')
        pix_color = 'FF%02X%02X%02X' % (pix[col - 1, row - 1][0], pix[col - 1, row - 1][1], pix[col - 1, row - 1][2])
        cell.style = Style(fill=PatternFill(patternType='solid', fgColor=Color(rgb=pix_color)))
    ws.row_dimensions[row].height = 6
for i in range(1, w):
    ws.column_dimensions[get_column_letter(i)].width = 1

wb.save(filename='myself.xlsx')
print('ok')
