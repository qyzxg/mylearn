#!/usr/bin/python
# -*- coding:utf-8 -*-

from openpyxl import load_workbook

wb = load_workbook(filename='./text/test.xlsx')
ws = wb.worksheets[0]
rows = ws.rows
columns = ws.columns
content_r = []
content_c = []
for row in rows:
    line = [col.value for col in row]
    content_r.append(line)
print(content_r)

for col in columns:
    line = [row.value for row in col]
    content_c.append(line)
print(content_c)
print(ws.cell(row=2, column=2).value)
